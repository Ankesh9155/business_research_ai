"""
Generates test_research.xlsx for testing the /research/start API.

Sheets:
  - "Info" : Country | Industry | Job Title | Employee Size | Max Contacts
  - "TAL"  : Company | Domain
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill("solid", fgColor="2F5496")
DATA_FILL_A  = PatternFill("solid", fgColor="EBF0FA")
DATA_FILL_B  = PatternFill("solid", fgColor="FFFFFF")
BORDER_SIDE  = Side(style="thin", color="B0B8CC")
CELL_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)


def style_sheet(ws):
    """Apply header and alternating-row styles to a worksheet."""

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(
                horizontal="center" if row_idx == 1 else "left",
                vertical="center",
                wrap_text=True
            )
            if row_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.fill = DATA_FILL_A if row_idx % 2 == 0 else DATA_FILL_B

    # Auto column width
    for col in ws.columns:
        max_len    = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 4

    ws.row_dimensions[1].height = 28


def main():
    path = "test_research.xlsx"

    info_data = {
        "Country":       ["United States, India, United Kingdom"],
        "Industry":      ["Technology, SaaS, Fintech"],
        "Job Title":     ["CEO, CTO, VP of Engineering, Head of Product"],
        "Employee Size": ["50-200"],
        "Max Contacts":  [3],
    }

    tal_data = {
        "Company": [
            "Stripe", "Notion", "Figma", "Linear", "Vercel",
            "Rippling", "Gusto", "Brex", "Ramp", "Lattice",
        ],
        "Domain": [
            "stripe.com", "notion.so", "figma.com", "linear.app", "vercel.com",
            "rippling.com", "gusto.com", "brex.com", "ramp.com", "lattice.com",
        ],
    }

    info_df = pd.DataFrame(info_data)
    tal_df  = pd.DataFrame(tal_data)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        info_df.to_excel(writer, sheet_name="Info", index=False)
        tal_df.to_excel(writer, sheet_name="TAL",  index=False)

        style_sheet(writer.sheets["Info"])
        style_sheet(writer.sheets["TAL"])

    print(f"Created: {path}")
    print(f"  Info sheet — {len(info_df)} row(s), columns: {list(info_df.columns)}")
    print(f"  TAL  sheet — {len(tal_df)} companies, columns: {list(tal_df.columns)}")
    print()
    print("Test with:")
    print('  curl -X POST http://localhost:8000/research/start -F "file=@test_research.xlsx"')
    print("  or via Swagger UI at http://localhost:8000/docs")


if __name__ == "__main__":
    main()
